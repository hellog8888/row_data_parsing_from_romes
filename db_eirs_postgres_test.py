import io
import glob
import datetime
import openpyxl
import warnings
import pandas as pd
import psycopg2.extras
from psycopg2 import Error
from sqlalchemy import create_engine


warnings.simplefilter("ignore")

dict_for_operator = \
    {
        'Общество с ограниченной ответственностью «Скартел»': 'Скартел',
        'Общество с ограниченной ответственностью \"Скартел\"': 'Скартел',

        'Общество с ограниченной ответственностью \"Т2 Мобайл\"': 'Т2 Мобайл',
        'Общество с ограниченной ответственностью «Т2 Мобайл»': 'Т2 Мобайл',

        'Публичное акционерное общество «Мобильные ТелеСистемы»': 'МТС',
        'Публичное акционерное общество \"Мобильные ТелеСистемы\"': 'МТС',

        'Публичное акционерное общество \"МегаФон\"': 'МегаФон',
        'Публичное акционерное общество «МегаФон»': 'МегаФон',

        'Публичное акционерное общество \"Ростелеком\"': 'Ростелеком',
        'Публичное акционерное общество «Ростелеком»': 'Ростелеком',
        'Публичное акционерное общество междугородной и международной электрической связи \"Ростелеком\"': 'Ростелеком',

        'Публичное акционерное общество «Вымпел-Коммуникации»': 'ВымпелКом',
        'Публичное акционерное общество \"Вымпел-Коммуникации\"': 'ВымпелКом'
    }


dict_ETC = \
    {
        '18.1.1.3.': 'GSM',
        '18.1.1.8.': 'GSM',
        '18.1.1.5.': 'UMTS',
        '18.1.1.6.': 'UMTS',
        '18.7.1.': 'LTE',
        '18.7.4.': 'LTE',
        '18.7.5.': '5G NR',
        '19.2.': 'РРС'
    }


def to_sql(file):

    df = pd.read_excel(file).loc[:,
               ['Наименование РЭС', 'Адрес', '№ вида ЕТС', 'Владелец', 'Широта', 'Долгота', 'Частоты',
                'Дополнительные параметры', 'Классы излучения', 'Серия последнего действующего РЗ/СоР',
                'Номер последнего действующего РЗ/СоР']]

    df['№ вида ЕТС'] = [dict_ETC[x.strip()] for x in df['№ вида ЕТС']]
    df['Владелец'] = [dict_for_operator[x.strip()] for x in df['Владелец']]

    df['Серия_Номер_последнего_действующего_РЗ_СоР'] = df['Серия последнего действующего РЗ/СоР'].astype(str) + ' ' + df['Номер последнего действующего РЗ/СоР'].astype(str)

    df = df.drop(['Серия последнего действующего РЗ/СоР'], axis=1)
    df = df.drop(['Номер последнего действующего РЗ/СоР'], axis=1)

    print('data_reading_complite')

    try:
        engine = create_engine('postgresql://postgres:1234@localhost:5432/eirs')
        df.to_sql('cellular', engine, index=False)
        print("Table created successfully")

    except (Exception, Error) as error:
        print("Error while creating the table", error)


def convert_exel_to_csv(file):
    cur_time = datetime.datetime.now()
    time_now = f'{cur_time.day}-{cur_time.month:02}-{cur_time.year}_{cur_time.hour:02}_{cur_time.minute:02}_{cur_time.second:02}'

    file_all = pd.read_excel(file).loc[:,['Наименование РЭС', 'Адрес', '№ вида ЕТС', 'Владелец', 'Широта', 'Долгота', 'Частоты', 'Дополнительные параметры', 'Классы излучения', 'Серия последнего действующего РЗ/СоР', 'Номер последнего действующего РЗ/СоР']]

    file_all['Наименование РЭС'] = file_all['Наименование РЭС'].str.strip()
    file_all['Адрес'] = file_all['Адрес'].str.strip().str.replace('., ', ', ').str.replace('\"', '')
    file_all['№ вида ЕТС'] = [dict_ETC[x.strip()] for x in file_all['№ вида ЕТС']]
    file_all['Владелец'] = [dict_for_operator[x.strip()] for x in file_all['Владелец']]
    file_all['Широта'] = file_all['Широта'].str.strip()
    file_all['Долгота'] = file_all['Долгота'].str.strip()
    file_all['Частоты'] = file_all['Частоты'].str.strip()
    file_all['Дополнительные параметры'] = file_all['Дополнительные параметры'].str.strip()
    file_all['Классы излучения'] = file_all['Классы излучения'].str.strip()
    file_all['Серия последнего действующего РЗ/СоР'] = file_all['Серия последнего действующего РЗ/СоР'].str.strip()

    file_all.to_csv(f'source_folder\\{time_now}.csv', sep='^', index=False)


def write_to_postgres(file_csv):

    hostname = 'localhost'
    database = 'eirs'
    username = 'postgres'
    password = '1234'
    port_id = 5432
    connection = None

    try:
        with psycopg2.connect(host=hostname, dbname=database, user=username, password=password, port=port_id) as connection:
            with connection.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:

                cur.execute('DROP TABLE IF EXISTS cellular')
                create_table = """ CREATE TABLE IF NOT EXISTS cellular (
                                        РЭС                       text,
                                        Адрес                     varchar(230),
                                        ТИП_РЭС                   text,
                                        Владелец                  varchar(11),
                                        Широта                    varchar(9),
                                        Долгота                   varchar(9),
                                        Частоты                   varchar(756),
                                        Дополнительные_параметры  text,
                                        Классы_излучения          varchar(53),
                                        Серия_Номер_РЗ_СоР        text)
                                """
                cur.execute(create_table)

                with io.open(file_csv, mode="r", encoding='utf-8') as csv_file:
                    data = []
                    for row in csv_file.readlines():
                        data = row.strip().split('^')

                        cur.execute("INSERT INTO cellular (РЭС, Адрес, ТИП_РЭС, Владелец, Широта, Долгота, Частоты, Дополнительные_параметры, Классы_излучения, Серия_Номер_РЗ_СоР) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                         (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], f'{data[9]} {data[10]}'))

                    data.clear()

                connection.commit()

    except Exception as error:
        print(error)
    finally:
        if connection is not None:
            connection.close()


def convert_to_postgres(file_open):

    hostname = 'localhost'
    database = 'eirs'
    username = 'postgres'
    pwd = '1234'
    port_id = 5432
    conn = None

    try:
        with psycopg2.connect(
                host=hostname,
                dbname=database,
                user=username,
                password=pwd,
                port=port_id
        ) as conn:

            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:

                # удалить таблицу (не считать ошибкой)
                cur.execute('DROP TABLE IF EXISTS cellular')

                create_script = """ CREATE TABLE IF NOT EXISTS cellular (
                                        РЭС                       varchar(51),
                                        Адрес                     varchar(230),
                                        ТИП_РЭС                   varchar(5),
                                        Владелец                  varchar(11),
                                        Широта                    varchar(9),
                                        Долгота                   varchar(9),
                                        Частоты                   varchar(756),
                                        Дополнительные_параметры  varchar(590),
                                        Классы_излучения          varchar(53),
                                        Серия_Номер_РЗ_СоР        varchar(13))
                                """

                cur.execute(create_script)

                file_to_read = openpyxl.load_workbook(file_open, data_only=True)
                sheet = file_to_read['SQL Results']

                for row in range(2, sheet.max_row + 1):
                    data = []

                    for col in range(1, sheet.max_column + 1):
                        value = sheet.cell(row, col).value
                        data.append(value)

                    cur.execute(
                         "INSERT INTO cellular (РЭС, Адрес, ТИП_РЭС, Владелец, Широта, Долгота, Частоты, Дополнительные_параметры, Классы_излучения, Серия_Номер_РЗ_СоР) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                         (str(data[1]), str(data[2]), str(dict_ETC[data[3]]), str(dict_for_operator[data[6]]),
                          str(data[7]), str(data[8]), str(data[10]), str(data[11]), str(data[17]), f'{data[18]} {data[19]}'))

                    #print(data[1], data[2], dict_ETC[data[3]], dict_for_operator[data[6]], data[7], data[8], data[10], data[11], data[17], f'{data[18]} {data[19]}')
                    data.clear()

                conn.commit()

    except Exception as error:
        print(error)
    finally:
        if conn is not None:
            conn.close()